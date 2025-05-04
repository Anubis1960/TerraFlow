from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.utils.dataframe import dataframe_to_rows


def create_bar_chart(ws: Workbook, title: str, y_axis_title: str, x_axis_title: str, start_col: str, start_row: int) \
        -> None:
    chart = BarChart()
    chart.title = title
    chart.y_axis.title = y_axis_title
    chart.x_axis.title = x_axis_title
    chart.style = 13  # Apply a professional style
    chart.legend.position = "b"  # Place legend at the bottom
    chart.x_axis.majorGridlines = None  # Remove major gridlines for clarity

    _data = Reference(ws, min_col=2, min_row=1, max_row=13)
    categories = Reference(ws, min_col=1, min_row=2, max_row=13)

    chart.add_data(_data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, f"{start_col}{start_row}")


def create_line_chart(ws: Workbook, title: str, y_axis_title: str, x_axis_title: str, start_col: str, start_row: int) \
        -> None:
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = y_axis_title
    chart.x_axis.title = x_axis_title
    chart.style = 12  # Apply a professional style
    chart.legend.position = "b"

    _data = Reference(ws, min_col=2, min_row=1, max_row=13)
    categories = Reference(ws, min_col=1, min_row=2, max_row=13)

    chart.add_data(_data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, f"{start_col}{start_row}")


def export_to_excel(_data: dict) -> BytesIO:
    MONTH_MAP = {
        "01": "January", "02": "February", "03": "March", "04": "April",
        "05": "May", "06": "June", "07": "July", "08": "August",
        "09": "September", "10": "October", "11": "November", "12": "December"
    }

    # Convert records to DataFrame
    records_df = pd.DataFrame(
        [{**record["sensor_data"], "Timestamp": record["timestamp"]} for record in _data["record"]]
    )

    # Convert water usage to DataFrame
    water_usage_df = pd.DataFrame(_data["water_usage"])
    water_usage_df["Year"] = water_usage_df["date"].str[:4]
    water_usage_df["Month"] = water_usage_df["date"].str[5:].map(MONTH_MAP)

    # Create Excel workbook
    wb = Workbook()
    ws_records = wb.active
    ws_records.title = "Sensor Records"

    # Write records data to Excel
    for row in dataframe_to_rows(records_df, index=False, header=True):
        ws_records.append(row)

    # Write water usage data to Excel
    ws_water = wb.create_sheet("Water Usage")
    for row in dataframe_to_rows(water_usage_df.drop(columns=["Year"]), index=False, header=True):
        ws_water.append(row)

    # Ensure all months are included with zero values if missing
    all_months = pd.DataFrame({"Month": list(MONTH_MAP.values())})
    water_usage_df = (
        all_months.merge(water_usage_df, on="Month", how="left")
        .fillna({"water_used": 0}))

    # Pivot data to get yearly structure
    water_usage_pivot = water_usage_df.pivot(index="Month", columns="Year", values="water_used").fillna(0).reset_index()
    water_usage_pivot["MonthOrder"] = water_usage_pivot["Month"].map({v: k for k, v in MONTH_MAP.items()})
    water_usage_pivot = water_usage_pivot.sort_values("MonthOrder").drop(columns=["MonthOrder"])

    # Write yearly water usage data and create charts
    for year in water_usage_df["Year"].dropna().unique():
        ws_year = wb.create_sheet(year)
        year_data = water_usage_pivot[["Month", year]].fillna(0)

        for row in dataframe_to_rows(year_data, index=False, header=True):
            ws_year.append(row)

        create_bar_chart(ws_year, f"Monthly Water Usage for {year}", "Water Used", "Month", "D", 2)
        create_line_chart(ws_year, f"Monthly Trend for {year}", "Water Used", "Month", "H", 2)

    # Save workbook to BytesIO buffer
    buff = BytesIO()
    wb.save(buff)
    buff.seek(0)
    return buff