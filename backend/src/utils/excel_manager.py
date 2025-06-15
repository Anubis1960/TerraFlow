from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.utils.dataframe import dataframe_to_rows
import regex as re


def sanitize_sheet_title(title: str) -> str:
    invalid_chars = r'[\\/:*?"<>|]'  # Regular expression pattern for invalid characters
    sanitized = re.sub(invalid_chars, '', title)
    if len(title) > 31:  # Max length for sheet names is 31 characters
        sanitized = sanitized[:31]
    return sanitized.strip()


def create_bar_chart(ws: Workbook, title: str, y_axis_title: str, x_axis_title: str, start_col: str, start_row: int) \
        -> None:
    """
    Creates a bar chart in the specified worksheet with the given titles and starting position.

    :param ws: Workbook: The worksheet where the chart will be added.
    :param title: str: The title of the chart.
    :param y_axis_title: str: The title for the Y-axis.
    :param x_axis_title: str: The title for the X-axis.
    :param start_col: str: The column letter where the chart will start.
    :param start_row: int: The row number where the chart will start.
    :return: None
    """
    chart = BarChart()
    title = sanitize_sheet_title(title)
    chart.title = title
    chart.y_axis.title = y_axis_title
    chart.x_axis.title = x_axis_title
    chart.style = 13  # Apply a professional style
    chart.legend.position = "b"  # Place legend at the bottom
    chart.x_axis.majorGridlines = None  # Remove major gridlines for clarity

    _data = Reference(ws, min_col=2, min_row=1, max_row=13)
    categories = Reference(ws, min_col=1, min_row=2, max_row=13)

    chart.add_data(_data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, f"{start_col}{start_row}")


def create_line_chart(ws: Workbook, title: str, y_axis_title: str, x_axis_title: str, start_col: str, start_row: int) \
        -> None:
    """
    Creates a line chart in the specified worksheet with the given titles and starting position.

    :param ws: The worksheet where the chart will be added.
    :param title: The title of the chart.
    :param y_axis_title: The title for the Y-axis.
    :param x_axis_title: The title for the X-axis.
    :param start_col: The column letter where the chart will start.
    :param start_row: The row number where the chart will start.
    :return: None
    """
    chart = LineChart()
    title = sanitize_sheet_title(title)
    chart.title = title
    chart.y_axis.title = y_axis_title
    chart.x_axis.title = x_axis_title
    chart.style = 12  # Apply a professional style
    chart.legend.position = "b"

    _data = Reference(ws, min_col=2, min_row=1, max_row=13)
    categories = Reference(ws, min_col=1, min_row=2, max_row=13)

    chart.add_data(_data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, f"{start_col}{start_row}")


def export_to_excel(_data: dict) -> BytesIO:
    """
    Exports data to an Excel file with multiple sheets for sensor records and water usage.

    :param _data: The data containing sensor records and water usage.
    :return: BytesIO buffer containing the Excel file.
    """
    wb = Workbook()
    try:
        build_excel_sheet(wb, _data)
    except ValueError as e:
        print(f"Error building Excel sheet: {e}")
        return BytesIO()

    # Save workbook to BytesIO buffer
    buff = BytesIO()
    wb.save(buff)
    buff.seek(0)
    return buff


def build_excel_sheet(wb: Workbook, data, prefix='') -> None:
    """
    Builds an Excel sheet from the provided data, creating separate sheets for sensor records and water usage.

    :param wb: The workbook to which the sheets will be added.
    :param data: The data containing sensor records and water usage.
    :param prefix: A prefix for the sheet names.
    :return: None
    """
    MONTH_MAP = {
        "01": "January", "02": "February", "03": "March", "04": "April",
        "05": "May", "06": "June", "07": "July", "08": "August",
        "09": "September", "10": "October", "11": "November", "12": "December"
    }

    # Convert records to DataFrame

    if "record" not in data:
        raise ValueError("Data must contain 'record'.")

    records_df = pd.DataFrame(
        [{**record["sensor_data"], "Timestamp": record["timestamp"]} for record in data["record"]]
    )

    # Convert water usage to DataFrame


    # Create Excel workbook
    ws_records = wb.active
    title = sanitize_sheet_title(f"{prefix} Sensor Records")
    ws_records.title = title

    # Write records data to Excel
    for row in dataframe_to_rows(records_df, index=False, header=True):
        ws_records.append(row)

    if "water_usage" not in data:
        return

    if data["water_usage"] is None:
        print("No water usage data found.")
        return

    if len(data["water_usage"]) == 0:
        print("Water usage list is empty. Skipping processing.")
        return

    water_usage_df = pd.DataFrame(data["water_usage"])
    water_usage_df["Year"] = water_usage_df["date"].str[:4]
    water_usage_df["Month"] = water_usage_df["date"].str[5:7].map(MONTH_MAP)

    # Write water usage data to Excel
    title = sanitize_sheet_title(f"{prefix} Water Usage")

    ws_water = wb.create_sheet(title)
    for row in dataframe_to_rows(water_usage_df.drop(columns=["Year"]), index=False, header=True):
        ws_water.append(row)

    # Ensure all months are included with zero values if missing
    all_months = pd.DataFrame({"Month": list(MONTH_MAP.values())})
    water_usage_df = (
        all_months.merge(water_usage_df, on="Month", how="left")
        .fillna({"water_used": 0}))

    # Pivot data to get yearly structure
    water_usage_pivot = water_usage_df.pivot(index="Month", columns="Year", values="water_used").fillna(
        0).reset_index()
    water_usage_pivot["MonthOrder"] = water_usage_pivot["Month"].map({v: k for k, v in MONTH_MAP.items()})
    water_usage_pivot = water_usage_pivot.sort_values("MonthOrder").drop(columns=["MonthOrder"])

    # Write yearly water usage data and create charts
    for year in water_usage_df["Year"].dropna().unique():
        ws_year = wb.create_sheet(f"{prefix} {year}")
        year_data = water_usage_pivot[["Month", year]].fillna(0)

        for row in dataframe_to_rows(year_data, index=False, header=True):
            ws_year.append(row)

        create_bar_chart(ws_year, f"Monthly Water Usage for {year}", "Water Used", "Month", "D", 2)
        create_line_chart(ws_year, f"Monthly Trend for {year}", "Water Used", "Month", "H", 2)


def build_device_excel_sheet(ws, records_df: pd.DataFrame, water_usage_df: pd.DataFrame, device_name: str):
    """
    Writes sensor records and water usage to a single worksheet.
    Adds headers for separation.
    """
    ws.append([f"Device: {device_name}"])
    ws.append(["Sensor Records"])
    for row in dataframe_to_rows(records_df, index=False, header=True):
        ws.append(row)

    ws.append([])  # Blank line
    ws.append(["Water Usage"])
    for row in dataframe_to_rows(water_usage_df, index=False, header=True):
        ws.append(row)


def export_to_excel_devices(_data: list) -> BytesIO:
    """
    Exports a list of devices into an Excel file where each device has its own sheet,
    containing sensor records and water usage data.
    """
    wb = Workbook()
    # Remove default sheet created by Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    for idx, device in enumerate(_data):
        if 'record' not in device or 'water_usage' not in device:
            continue

        device_name = device.get("name", "Device")
        device_name = f"{device_name} {idx + 1}"
        print(f"Processing {device_name}")

        # Convert sensor records to DataFrame
        if not device["record"]:
            print(f"No records found for device {device_name}")
            continue
        try:
            records_df = pd.DataFrame([{**r["sensor_data"], "Timestamp": r["timestamp"]} for r in device["record"]])

            # Convert water usage to DataFrame
            water_usage_df = pd.DataFrame(device["water_usage"])
        except Exception as e:
            print(f"Error processing device {device_name}: {e}")
            continue

        # Create new sheet
        sheet_title = sanitize_sheet_title(device_name)
        ws = wb.create_sheet(sheet_title)

        # Write both datasets to this sheet
        build_device_excel_sheet(ws, records_df, water_usage_df, device_name)

    # Save workbook to buffer
    buff = BytesIO()
    wb.save(buff)
    buff.seek(0)
    return buff
